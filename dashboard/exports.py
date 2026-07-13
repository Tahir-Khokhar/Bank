# Enable postponed evaluation of type hints
from __future__ import annotations

# Import the CSV module to create CSV files
import csv

# Import the io module for in-memory file handling
import io

# Import HttpResponse to send downloadable files to the browser
from django.http import HttpResponse

# Import Django timezone utilities
from django.utils import timezone


# Create and return a downloadable CSV file
def csv_response(rows, name):

    # Create an HTTP response with CSV content type
    response = HttpResponse(content_type="text/csv")

    # Set the download file name
    response["Content-Disposition"] = (
        f'attachment; filename="{name}_{timezone.now():%Y%m%d}.csv"')

    # Create a CSV writer
    writer = csv.writer(response)

    # Write each row into the CSV file
    for row in rows:
        writer.writerow(row)

    # Return the completed response
    return response


# Create and return a downloadable Excel file
def excel_response(rows, name, title=None):

    # Import Workbook class for creating Excel files
    from openpyxl import Workbook

    # Import Excel cell styling classes
    from openpyxl.styles import Alignment, Font, PatternFill

    # Create a new workbook
    wb = Workbook()

    # Get the active worksheet
    ws = wb.active

    # Set the worksheet title
    ws.title = (name[:28] or "Report")

    # Create a dark background color for header cells
    header_fill = PatternFill("solid", fgColor="0A2540")

    # Start writing from the first row
    start = 1

    # Add the report title if provided
    if title:
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        start = 3

    # Write the header row
    for c_idx, value in enumerate(rows[0], start=1):

        # Create a cell for each header
        cell = ws.cell(row=start, column=c_idx, value=value)

        # Make header text bold and white
        cell.font = Font(bold=True, color="FFFFFF")

        # Apply background color
        cell.fill = header_fill

        # Center-align the header text
        cell.alignment = Alignment(horizontal="center")

    # Write the remaining data rows
    for r_offset, row in enumerate(rows[1:], start=start + 1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_offset, column=c_idx, value=value)

    # Automatically adjust column widths
    for col in ws.columns:

        # Find the longest value in the column
        width = max((len(str(c.value)) for c in col if c.value), default=10)

        # Set the column width
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 48)

    # Create an Excel file response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Set the download file name
    response["Content-Disposition"] = (
        f'attachment; filename="{name}_{timezone.now():%Y%m%d}.xlsx"')

    # Save the workbook into the response
    wb.save(response)

    # Return the completed response
    return response


# Create and return a downloadable PDF file
def pdf_response(rows, name, title="Report", subtitle=""):

    # Import ReportLab color utilities
    from reportlab.lib import colors

    # Import page size functions
    from reportlab.lib.pagesizes import A4, landscape

    # Import default text styles
    from reportlab.lib.styles import getSampleStyleSheet

    # Import millimeter measurement unit
    from reportlab.lib.units import mm

    # Import PDF layout components
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    )

    # Create an in-memory buffer
    buffer = io.BytesIO()

    # Create a PDF document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=18 * mm,
        bottomMargin=15 * mm,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
    )

    # Load default styles
    styles = getSampleStyleSheet()

    # Create the PDF content list
    elements = [Paragraph(f"<b>{title}</b>", styles["Title"])]

    # Add subtitle if provided
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))

    # Add the report generation time
    elements.append(
        Paragraph(
            f"Generated: {timezone.now():%Y-%m-%d %H:%M}",
            styles["Normal"],
        )
    )

    # Add spacing below the title
    elements.append(Spacer(1, 8 * mm))

    # Create a table from the data
    table = Table(rows, repeatRows=1)

    # Apply styles to the table
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0A2540")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F1F4F9")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D7DEE8")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    # Add the table to the document
    elements.append(table)

    # Generate the PDF
    doc.build(elements)

    # Read the PDF from memory
    pdf = buffer.getvalue()

    # Close the memory buffer
    buffer.close()

    # Create an HTTP response for the PDF
    response = HttpResponse(content_type="application/pdf")

    # Set the download file name
    response["Content-Disposition"] = (
        f'attachment; filename="{name}_{timezone.now():%Y%m%d}.pdf"')

    # Write the PDF into the response
    response.write(pdf)

    # Return the completed response
    return response


# Return the correct export file based on the selected format
def export_dispatch(fmt, rows, name, title="Report", subtitle=""):

    """Return the correct export response for csv, excel, or pdf."""

    # Export as Excel
    if fmt == "excel":
        return excel_response(rows, name, title=title)

    # Export as PDF
    if fmt == "pdf":
        return pdf_response(rows, name, title=title, subtitle=subtitle)

    # Export as CSV by default
    return csv_response(rows, name)
